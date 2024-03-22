#
# Name: excel2yaml.py
#
# Description: Reads information in an Excel file (based on "ARS Template.xlsx")
#              and creates a YAML file containing an ARS ReportingEvent. The
#              YAML file will be named according to the name of the input Excel file.
#              For example, an Excel file called "Study 1 CSR.xlsx" will generate a
#              YAML file called "Study 1 CSR.yaml" (in the same folder as the
#              original Excel file). The id, version and name of the ReportingEvent
#              will be as specified on the ReportEvent sheet of the Excel file.
#
# Usage: python excel2yaml.py -x <path to Excel file>
#
# Example: python excel2yaml.py -x '..\..\workfiles\examples\Sprint 12 Examples.xlsx'
#

import os
import sys
import argparse
import re
from linkml_runtime.dumpers import yaml_dumper
from openpyxl import load_workbook, Workbook

# Add top-level folder to path so that project folder can be found
lib_path = os.path.abspath(os.path.join(__file__, "..", "..", ".."))
if lib_path not in sys.path:
    sys.path.append(lib_path)

from project.ars_ldm import *


def parse_arguments():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-x", "--excel_file", help="Excel file containing details of reporting event"
    )
    args = parser.parse_args()
    return args


args = parse_arguments()

if not args.excel_file:
    print("The name of the Excel file must be specified.")
    exit(1)

filename = os.path.split(args.excel_file)[-1]

if "".join(filename.split(".")[-1]).lower() != "xlsx":
    print("Only xlsx files are valid.")
    exit(1)

rptevtfname = "".join(filename.split(".")[0])

wb: Workbook
wb = load_workbook(filename=args.excel_file)


def get_nlist(sheetname: str, clist: str, values: list[list]) -> NestedList:
    nlists = {}
    lvl = 0
    for value in values:
        if value[4] == 1:
            nlists[value[0]] = NestedList(
                listItems=[
                    OrderedListItem(
                        level=value[0],
                        name=value[1],
                        description=value[2],
                        label=value[3],
                        order=value[4],
                        analysisId=value[5],
                        outputId=value[6],
                    )
                ]
            )
            lvl = value[0]
        elif value[0] <= lvl:
            if value[0] < lvl:
                for _i in range(lvl)[: (value[0] - 1) : -1]:
                    nlists[_i].listItems[-1].sublist = nlists.pop(_i + 1)
                    lvl = value[0]
            nlists[value[0]].listItems.append(
                OrderedListItem(
                    level=value[0],
                    name=value[1],
                    description=value[2],
                    label=value[3],
                    order=value[4],
                    analysisId=value[5],
                    outputId=value[6],
                )
            )
        else:
            print(
                f"Unexpected entry in {sheetname} content list '{clist}': level={value[0]}, order={value[2]}"
            )
    else:
        for _i in range(lvl)[:0:-1]:
            nlists[_i].listItems[-1].sublist = nlists.pop(_i + 1)

    return nlists[1]


clists = {}
for sheetname in ["MainListOfContents", "OtherListsOfContents"]:
    wsLopx = wb[sheetname]
    for value in wsLopx.iter_rows(min_row=2, values_only=True):
        if all([c is None for c in value]):
            continue
        if value[0] in clists:
            clists[value[0]]["values"].append(value[3:])
            if value[1] != clists[value[0]]["description"]:
                print(
                    f"Inconsistent description found for {value[0]} "
                    + f"content list in {sheetname} sheet: '{value[1]}'"
                    + f"/'{clists[value[0]]['description']}'"
                )
            if value[2] != clists[value[0]]["label"]:
                print(
                    f"Inconsistent label found for {value[0]} "
                    + f"content list in {sheetname} sheet: '{value[2]}'"
                    + f"/'{clists[value[0]]['label']}'"
                )
        else:
            clists[value[0]] = {
                "sheetname": sheetname,
                "main": (sheetname == "MainListOfContents"),
                "description": value[1],
                "label": value[2],
                "values": [value[3:]],
            }

found_main: bool = False

for k, v in clists.items():
    if clists[k]["main"]:
        if found_main:
            print(
                f"Only one main list of contents allowed. Content list '{k}' ignored."
            )
            x = clists[k].pop("main")
            continue
        else:
            found_main = True
    clists[k]["clist"] = ListOfContents(
        name=k,
        description=v["description"],
        label=v["label"],
        contentsList=get_nlist(sheetname=v["sheetname"], clist=k, values=v["values"]),
    )


def get_content_lists(main: bool) -> list[ListOfContents]:
    return [v["clist"] for v in clists.values() if "main" in v and v["main"] == main]


wsRptEvt = wb["ReportingEvent"]
for value in wsRptEvt.iter_rows(min_row=2, max_row=2, values_only=True):
    if all([c is None for c in value]):
        continue
    rptevt = ReportingEvent(
        id=value[0],
        version=value[1],
        name=value[2],
        description=value[3],
        label=value[4],
        mainListOfContents=get_content_lists(main=True)[0],
        otherListsOfContents=get_content_lists(main=False),
    )


wsRefDoc = wb["ReferenceDocuments"]

for value in wsRefDoc.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue
    refdoc = ReferenceDocument(
        id=value[0],
        name=value[1],
        description=value[2],
        label=value[3],
        location=value[4],
    )
    rptevt.referenceDocuments.append(refdoc)

wsCat = wb["Categorizations"]

catnid = ""
cats = {}
catns = {}

for value in wsCat.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue
    if value[0] != catnid:
        if value[2] is None:
            rptevt.analysisOutputCategorizations.append(
                AnalysisOutputCategorization(
                    id=value[0],
                    label=value[1],
                    categories=[AnalysisOutputCategory(id=value[3], label=value[4])],
                )
            )
            catns[value[0]] = rptevt.analysisOutputCategorizations[-1]
        else:
            if value[2] in cats:
                cats[value[2]].subCategorizations.append(
                    AnalysisOutputCategorization(
                        id=value[0],
                        label=value[1],
                        categories=[
                            AnalysisOutputCategory(id=value[3], label=value[4])
                        ],
                    )
                )
                catns[value[0]] = cats[value[2]].subCategorizations[-1]
            else:
                print(f"Unknown parent_category_id: {value[2]}")

        catnid = value[0]
        pcatid = value[2]

    else:

        catns[value[0]].categories.append(
            AnalysisOutputCategory(id=value[3], label=value[4])
        )
        if value[2] != pcatid:
            print(
                f"Inconsistent parent_category_id values defined for categorization {value[0]}: {pcatid} / {value[2]}"
            )

    cats[value[3]] = catns[value[0]].categories[-1]

wsGDS = wb["GlobalDisplaySections"]
sidx = {}
dsubsects = {}
for value in wsGDS.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue
    ssect = DisplaySubSection(id=value[1], text=value[2])

    if value[0] in sidx:
        sidx[value[0]].subSections.append(ssect)
    else:
        rptevt.globalDisplaySections.append(
            GlobalDisplaySection(sectionType=value[0], subSections=[ssect])
        )
        sidx[value[0]] = rptevt.globalDisplaySections[-1]

    dsubsects[value[1]] = sidx[value[0]].subSections[-1]

wsTermEx = wb["TerminologyExtensions"]
termexidx = {}
spterms = {}
for value in wsTermEx.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue

    spterm = SponsorTerm(id=value[2], submissionValue=value[3], description=value[4])

    if value[0] in termexidx:
        termexidx[value[0]].sponsorTerms.append(spterm)
        if value[1] != termexidx[value[0]].enumeration:
            print(
                f"More than one enumeration specified for terminology extension {value[0]}: {termexidx[value[0]].enumeration} / {value[1]}"
            )
    else:
        rptevt.terminologyExtensions.append(
            TerminologyExtension(
                id=value[0], enumeration=value[1], sponsorTerms=[spterm]
            )
        )
        termexidx[value[0]] = rptevt.terminologyExtensions[-1]

    spterms[value[1]] = termexidx[value[0]].sponsorTerms[-1]

wsAnSet = wb["AnalysisSets"]

asid = ""
anset: AnalysisSet = None
wcs = {}

for value in wsAnSet.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue

    if value[6] is None:
        if value[9] is not None:
            wc = WhereClause(
                level=value[4],
                order=value[5],
                condition=WhereClauseCondition(
                    dataset=value[8],
                    variable=value[9],
                    comparator=value[10],
                    value=str(value[11]).split(" | ")
                    if value[11] is not None and str(value[11]).find(" | ") > -1
                    else value[11],
                )
                if value[9] is not None
                else None,
            )
        else:
            wc = ReferencedAnalysisSet(
                level=value[4], order=value[5], subClauseId=value[7]
            )
    else:
        wc = WhereClause(
            level=value[4],
            order=value[5],
            compoundExpression=CompoundSetExpression(
                logicalOperator=value[6], whereClauses=[]
            )
            if value[6] is not None
            else None,
        )
        wcs[value[4]] = wc

    if value[0] != asid:
        if asid != "":
            rptevt.analysisSets.append(anset)
        anset = AnalysisSet(
            id=value[0],
            name=value[1],
            description=value[2],
            label=value[3],
            level=wc.level,
            order=wc.order,
            condition=wc.condition,
            compoundExpression=wc.compoundExpression,
        )
        asid = value[0]
    else:
        if isinstance(wcs[value[4] - 1].compoundExpression.whereClauses, list):
            wcs[value[4] - 1].compoundExpression.whereClauses.extend([wc])
        else:
            wcs[value[4] - 1].compoundExpression.whereClauses = [wc]
else:
    rptevt.analysisSets.append(anset)

wsAnGrp = wb["AnalysisGroupings"]
grpngid = ""
grpid = ""

grpng = GroupingFactor(id="dummy", dataDriven=False, name="dummy")
for value in wsAnGrp.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue

    if str(value[0]) != grpngid:
        wcs = {}

    if value[7] is not None:
        if value[13] is None:
            if value[16] is not None:
                wc = WhereClause(
                    level=value[11],
                    order=value[12],
                    condition=WhereClauseCondition(
                        dataset=value[15],
                        variable=value[16],
                        comparator=value[17],
                        value=str(value[18]).split(" | ")
                        if value[18] is not None and str(value[17]).find(" | ") > -1
                        else value[18],
                    )
                    if value[16] is not None
                    else None,
                )
            else:
                wc = ReferencedGroup(
                    level=value[11], order=value[12], subClauseId=value[14]
                )
        else:
            wc = WhereClause(
                level=value[11],
                order=value[12],
                compoundExpression=CompoundGroupExpression(
                    logicalOperator=value[13], whereClauses=[]
                )
                if value[13] is not None
                else None,
            )
            wcs[value[11]] = wc

        if value[7] != grpid:
            if grpid != "":
                grpng.groups.append(grp)
            grp = Group(
                id=value[7],
                name=value[8],
                description=value[9],
                label=value[10],
                level=wc.level,
                order=wc.order,
                condition=wc.condition,
                compoundExpression=wc.compoundExpression,
            )
            grpid = value[7]
        else:
            if isinstance(wcs[value[11] - 1].compoundExpression.whereClauses, list):
                wcs[value[11] - 1].compoundExpression.whereClauses.extend([wc])
            else:
                wcs[value[11] - 1].compoundExpression.whereClauses = [wc]
    else:
        if grpid != "":
            grpng.groups.append(grp)
        wc = None
        grpid = ""

    # If this is a new grouping...
    if str(value[0]) != grpngid:
        # Store what's been built for the previous grouping
        if grpng.id != "dummy":
            rptevt.analysisGroupings.append(grpng)
        grpngid = str(value[0])
        # Then process the new grouping
        grpng = GroupingFactor(
            id=grpngid,
            name=value[1],
            description=value[2],
            label=value[3],
            groupingDataset=value[4],
            groupingVariable=value[5],
            dataDriven=value[6],
            groups=[],
        )
#    else:
#        grpng.groups.append(DataGroup(id=value[5],label=value[6],level=value[7],order=value[8],condition=WhereClauseCondition(dataset=value[10],variable=value[11],comparator=value[12],value=str(value[13]).split(" | ") if value[13] is not None and str(value[13]).find(" | ") > -1 else value[13])))
else:
    if grpid != "":
        grpng.groups.append(grp)
        rptevt.analysisGroupings.append(grpng)

wsDss = wb["DataSubsets"]

dss: DataSubset = None
dssid = ""
wcs = {}

for value in wsDss.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue

    if value[6] is None:
        if value[9] is not None:
            wc = WhereClause(
                level=value[4],
                order=value[5],
                condition=WhereClauseCondition(
                    dataset=value[8],
                    variable=value[9],
                    comparator=value[10],
                    value=str(value[11]).split(" | ")
                    if value[11] is not None and str(value[11]).find(" | ") > -1
                    else value[11],
                )
                if value[9] is not None
                else None,
            )
        else:
            wc = ReferencedDataSubset(
                level=value[4], order=value[5], subClauseId=value[7]
            )
    else:
        wc = WhereClause(
            level=value[4],
            order=value[5],
            compoundExpression=CompoundSubsetExpression(
                logicalOperator=value[6],
                whereClauses=[],
            )
            if value[6] is not None
            else None,
        )
        wcs[value[4]] = wc

    if value[0] != dssid:
        if dssid != "":
            rptevt.dataSubsets.append(dss)
        dss = DataSubset(
            id=value[0],
            name=value[1],
            description=value[2],
            label=value[3],
            level=wc.level,
            order=wc.order,
            condition=wc.condition,
            compoundExpression=wc.compoundExpression,
        )
        dssid = value[0]
    else:
        if isinstance(wcs[value[4] - 1].compoundExpression.whereClauses, list):
            wcs[value[4] - 1].compoundExpression.whereClauses.extend([wc])
        else:
            wcs[value[4] - 1].compoundExpression.whereClauses = [wc]
else:
    if dss:
        rptevt.dataSubsets.append(dss)

results = {}

wsRslts = wb["AnalysisResults"]
anid = ""
for value in wsRslts.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue
    if str(value[0]) != anid:
        results[value[0]] = [
            OperationResult(
                operationId=value[4],
                rawValue=value[19],
                formattedValue=value[20],
                resultGroups=[
                    ResultGroup(
                        groupingId=value[x], groupId=value[y], groupValue=value[z]
                    )
                    for x, y, z in [[7, 8, 10], [11, 12, 14], [15, 16, 18]]
                    if value[x] is not None
                ],
            )
        ]
        anid = str(value[0])
    else:
        results[value[0]].append(
            OperationResult(
                operationId=value[4],
                rawValue=value[19],
                formattedValue=value[20],
                resultGroups=[
                    ResultGroup(
                        groupingId=value[x], groupId=value[y], groupValue=value[z]
                    )
                    for x, y, z in [[7, 8, 10], [11, 12, 14], [15, 16, 18]]
                    if value[x] is not None
                ],
            )
        )


def get_docrefs(sheetname: str) -> dict:

    wsDocRef = wb[sheetname]

    docrefs = {
        "sheetname": sheetname,
        "Documentation": dict(),
        "ProgrammingCode": dict(),
    }

    docidx = {
        "sheetname": sheetname,
        "Documentation": dict(),
        "ProgrammingCode": dict(),
    }

    for value in wsDocRef.iter_rows(min_row=2, values_only=True):
        if all([c is None for c in value]):
            continue

        pageref = None

        if value[3] is not None:
            if value[3] == "NamedDestination":
                pageref = PageNameRef(
                    refType=value[3],
                    label=value[4],
                    pageNames=value[5].split("|")
                    if value[5] is not None and str(value[5]).find("|") > -1
                    else str(value[5])
                    if value[5]
                    else None,
                )
            elif value[3] == "PhysicalRef":
                if re.search("^[0-9]+-[0-9]+$", str(value[5])):
                    pageref = PageNumberRangeRef(
                        refType=value[3],
                        label=value[4],
                        firstPage=value[5].split("-")[0],
                        lastPage=value[5].split("-")[1],
                    )
                elif re.search("^[0-9]+(\|[0-9]+)*$", str(value[5])):
                    pageref = PageNumberListRef(
                        refType=value[3],
                        label=value[4],
                        pageNumbers=value[5].split("|")
                        if value[5] is not None and str(value[5]).find("|") > -1
                        else str(value[5])
                        if value[5]
                        else None,
                    )
                else:
                    print(
                        f"Invalid pageRef_pages value on {sheetname} sheet: {value[5]}"
                    )
            else:
                print(f"Invalid pageRef_refType value on {sheetname} sheet: {value[5]}")

        if value[0] in docrefs[value[1]]:
            if value[2] not in docidx[value[1]][value[0]]:
                docrefs[value[1]][value[0]].append(
                    DocumentReference(referenceDocumentId=value[2])
                )
                docidx[value[1]][value[0]][value[2]] = docrefs[value[1]][value[0]][-1]
                docidx[value[1]][value[0]][value[2]].pageRefs = (
                    [pageref] if pageref else None
                )
            elif pageref:
                docidx[value[1]][value[0]][value[2]].pageRefs.append(pageref)
        else:
            docrefs[value[1]][value[0]] = [
                DocumentReference(referenceDocumentId=value[2])
            ]
            docidx[value[1]][value[0]] = {value[2]: docrefs[value[1]][value[0]][-1]}
            docidx[value[1]][value[0]][value[2]].pageRefs = (
                [pageref] if pageref else None
            )

    return docrefs


def get_params(sheetname: str, template: bool = False) -> dict:

    wsParams = wb[sheetname]

    params = {}

    for value in wsParams.iter_rows(min_row=2, values_only=True):
        if all([c is None for c in value]):
            continue

        if template:
            param = TemplateCodeParameter(
                name=value[1],
                description=value[2],
                label=value[3],
                valueSource=value[4],
                value=value[5].split("|")
                if value[5] is not None and str(value[5]).find("|") > -1
                else str(value[5])
                if value[5]
                else None,
            )
        else:
            param = AnalysisOutputCodeParameter(
                name=value[1], description=value[2], label=value[3], value=value[4]
            )

        if value[0] in params:
            params[value[0]].append(param)
        else:
            params[value[0]] = [param]

    return params


def get_progcode(
    sheetname: str, docrefs: dict, params: dict, template: bool = False
) -> dict:

    wsProgCode = wb[sheetname]

    progcode = {}

    for value in wsProgCode.iter_rows(min_row=2, values_only=True):
        if all([c is None for c in value]):
            continue

        if value[2] == "Code":
            if template:
                progcode[value[0]] = AnalysisProgrammingCodeTemplate(
                    context=value[1],
                    code=value[3],
                    parameters=params[value[0]] if value[0] in params else None,
                )
            else:
                progcode[value[0]] = AnalysisOutputProgrammingCode(
                    context=value[1],
                    code=value[3],
                    parameters=params[value[0]] if value[0] in params else None,
                )
        elif value[2] == "DocumentRef":
            if value[0] in docrefs["ProgrammingCode"]:
                if template:
                    progcode[value[0]] = AnalysisProgrammingCodeTemplate(
                        context=value[1],
                        documentRef=docrefs["ProgrammingCode"][value[0]][0],
                        parameters=params[value[0]] if value[0] in params else None,
                    )
                else:
                    progcode[value[0]] = AnalysisOutputProgrammingCode(
                        context=value[1],
                        documentRef=docrefs["ProgrammingCode"][value[0]][0],
                        parameters=params[value[0]] if value[0] in params else None,
                    )
            else:
                print(
                    f"Programming code specified as DocumentRef for {value[0]} on {sheetname} sheet, but no matching ProgrammingCode DocumentRef was found on the {docrefs['sheetname']} sheet."
                )
        else:
            print(f"Invalid specifiedAs value on the {sheetname} sheet: {value[2]}")

    return progcode


mparams = get_params(sheetname="AnalysisMethodCodeParameters", template=True)
mdocrefs = get_docrefs(sheetname="AnalysisMethodDocumentRefs")
mprogcode = get_progcode(
    sheetname="AnalysisMethodCodeTemplate",
    docrefs=mdocrefs,
    params=mparams,
    template=True,
)

wsMth = wb["AnalysisMethods"]
mthid = ""
method = AnalysisMethod(
    id="dummy", name="dummy", operations=[Operation(id="dummy", name="dummy", order=0)]
)
for value in wsMth.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue
    # If this is a new method...
    if str(value[0]) != mthid:
        # Store what's been built for the previous method
        if method.id != "dummy":
            rptevt.methods.append(method)
        mthid = str(value[0])
        # Then process the new method
        method = AnalysisMethod(
            id=mthid,
            name=value[1],
            description=value[2],
            label=value[3],
            operations=[
                Operation(
                    id=value[4],
                    name=value[5],
                    description=value[6],
                    label=value[7],
                    order=value[8],
                    resultPattern=value[9],
                    referencedOperationRelationships=[
                        ReferencedOperationRelationship(
                            id=value[10],
                            referencedOperationRole=OperationRole(
                                controlledTerm=value[11]
                            )
                            if value[11] in OperationRoleEnum
                            else SponsorOperationRole(sponsorTermId=value[11]),
                            operationId=value[12],
                            analysisId=value[13],
                            description=value[14],
                        ),
                        ReferencedOperationRelationship(
                            id=value[15],
                            referencedOperationRole=OperationRole(
                                controlledTerm=value[16]
                            )
                            if value[16] in OperationRoleEnum
                            else SponsorOperationRole(sponsorTermId=value[16]),
                            operationId=value[17],
                            analysisId=value[18],
                            description=value[19],
                        ),
                    ]
                    if value[10] is not None
                    else None,
                )
            ],
            documentRefs=mdocrefs["Documentation"][mthid]
            if mthid in mdocrefs["Documentation"]
            else None,
            codeTemplate=mprogcode[mthid] if mthid in mprogcode else None,
        )
    else:
        method.operations.append(
            Operation(
                id=value[4],
                name=value[5],
                description=value[6],
                label=value[7],
                order=value[8],
                resultPattern=value[9],
                referencedOperationRelationships=[
                    ReferencedOperationRelationship(
                        id=value[10],
                        referencedOperationRole=OperationRole(controlledTerm=value[11])
                        if value[11] in OperationRoleEnum
                        else SponsorOperationRole(sponsorTermId=value[11]),
                        operationId=value[12],
                        analysisId=value[13],
                        description=value[14],
                    ),
                    ReferencedOperationRelationship(
                        id=value[15],
                        referencedOperationRole=OperationRole(controlledTerm=value[16])
                        if value[16] in OperationRoleEnum
                        else SponsorOperationRole(sponsorTermId=value[16]),
                        operationId=value[17],
                        analysisId=value[18],
                        description=value[19],
                    ),
                ]
                if value[10] is not None
                else None,
            )
        )
else:
    rptevt.methods.append(method)

aparams = get_params(sheetname="AnalysisCodeParameters")
adocrefs = get_docrefs(sheetname="AnalysisDocumentRefs")
aprogcode = get_progcode(
    sheetname="AnalysisProgrammingCode", docrefs=adocrefs, params=aparams
)

wsAn = wb["Analyses"]

for value in wsAn.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue

    if value[6] is not None:
        if value[6] in AnalysisReasonEnum:
            anreas = AnalysisReason(controlledTerm=value[6])
        else:
            anreas = SponsorAnalysisReason(sponsorTermId=value[6])
    else:
        anreas = None

    if value[7] is not None:
        if value[7] in AnalysisPurposeEnum:
            anpurp = AnalysisPurpose(controlledTerm=value[7])
        else:
            anpurp = SponsorAnalysisPurpose(sponsorTermId=value[7])
    else:
        anpurp = None

    analysis = Analysis(
        id=value[0],
        version=value[1],
        name=value[2],
        description=value[3],
        label=value[4],
        reason=anreas,
        purpose=anpurp,
        documentRefs=adocrefs["Documentation"][value[0]]
        if value[0] in adocrefs["Documentation"]
        else None,
        analysisSetId=value[8],
        dataSubsetId=value[15],
        dataset=value[16],
        variable=value[17],
        methodId=value[18],
        referencedAnalysisOperations=[
            ReferencedAnalysisOperation(
                referencedOperationRelationshipId=value[x], analysisId=value[y]
            )
            for x, y in [[19, 20], [21, 22]]
            if value[x] is not None
        ],
        programmingCode=aprogcode[value[0]] if value[0] in aprogcode else None,
        results=results[value[0]] if value[0] in results else None,
    )
    if value[5] is not None:
        if value[5].find(" | ") > -1:
            for catid in value[5].split(" | "):
                if catid in cats:
                    analysis.categoryIds.extend([catid])
                else:
                    print(f"Unknown category_id for analysis {value[0]}: {catid}")
        else:
            if value[5] in cats:
                analysis.categoryIds.extend([value[5]])
            else:
                print(f"Unknown category_id for analysis {value[0]}: {value[5]}")
    grpord = 0
    for _i in [9, 11, 13]:
        if value[_i] is None:
            break
        grpord += 1
        analysis.orderedGroupings.append(
            OrderedGroupingFactor(
                order=grpord, groupingId=value[_i], resultsByGroup=value[_i + 1]
            )
        )
    rptevt.analyses.append(analysis)

wsOFile = wb["OutputFiles"]

ofiles = {}
for value in wsOFile.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue

    fs = OutputFile(
        name=value[1],
        description=value[2],
        label=value[3],
        location=value[4],
        fileType=OutputFileType(controlledTerm=value[5])
        if value[5] in OutputFileTypeEnum
        else SponsorOutputFileType(sponsorTermId=value[5]),
    )
    if value[0] in ofiles:
        ofiles[value[0]].append(fs)
    else:
        ofiles[value[0]] = [fs]

wsDisp = wb["Displays"]

displays = {}
dispid = ""
for value in wsDisp.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue

    if value[9] is None:
        ss = OrderedSubSectionRef(order=value[7], subSectionId=value[8])
        if value[8] not in dsubsects:
            print(
                f"Display Subsection id {value[8]} is referenced (without text), but this id has not been defined (with text)"
            )
    else:
        ss = OrderedSubSection(
            order=value[7], subSection=DisplaySubSection(id=value[8], text=value[9])
        )
        if value[8] in dsubsects:
            print(
                f"Duplicate display subsection id (with text defined) found: {value[8]}"
            )
        else:
            dsubsects[value[8]] = ss

    if str(value[0]) != dispid:
        if dispid != "":
            displays[dispid] = display
        display = OutputDisplay(
            id=value[0],
            name=value[1],
            description=value[2],
            label=value[3],
            version=value[4],
            displayTitle=value[5],
            displaySections=[
                DisplaySection(sectionType=value[6], orderedSubSections=[ss])
            ],
        )
        dispid = str(value[0])
        scttype = str(value[6])
    elif value[6] != scttype:
        display.displaySections.append(
            DisplaySection(sectionType=value[6], orderedSubSections=[ss])
        )
        scttype = value[6]
    else:
        display.displaySections[-1].orderedSubSections.append(ss)
else:
    displays[dispid] = display

oparams = get_params(sheetname="OutputCodeParameters")
odocrefs = get_docrefs(sheetname="OutputDocumentRefs")
oprogcode = get_progcode(
    sheetname="OutputProgrammingCode", docrefs=odocrefs, params=oparams
)

wsOutput = wb["Outputs"]

for value in wsOutput.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue
    output = Output(
        id=value[0],
        version=value[1],
        name=value[2],
        description=value[3],
        label=value[4],
        displays=[
            OrderedDisplay(order=ix + 1, display=displays[value[6 + ix]])
            for ix in range(2)
            if value[6 + ix]
        ],
        documentRefs=odocrefs["Documentation"][value[0]]
        if value[0] in odocrefs["Documentation"]
        else None,
        programmingCode=oprogcode[value[0]] if value[0] in oprogcode else None,
        fileSpecifications=ofiles[value[0]] if value[0] in ofiles else None,
    )
    if value[5] is not None:
        if value[5].find(" | ") > -1:
            for catid in value[5].split(" | "):
                if catid in cats:
                    output.categoryIds.extend([catid])
                else:
                    print(f"Unknown category_id for output {value[0]}: {catid}")
        else:
            if value[5] in cats:
                output.categoryIds.extend([value[5]])
            else:
                print(f"Unknown category_id for output {value[0]}: {value[5]}")
    rptevt.outputs.append(output)

yaml_str = yaml_dumper.dumps(rptevt)

with open(
    os.path.join("".join(os.path.split(args.excel_file)[0:-1]), rptevtfname + ".yaml"),
    "w",
    encoding="utf-8",
) as f:
    f.write(yaml_str)
